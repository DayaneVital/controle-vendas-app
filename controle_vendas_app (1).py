
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from datetime import date

st.set_page_config(page_title="Controle de Vendas - Loja de Roupas", layout="wide")
st.title("👚 Controle de Vendas - Loja de Roupas")

st.sidebar.header("🔁 Navegação")
aba = st.sidebar.radio("Escolha uma aba:", ["Cadastrar Venda", "Cadastrar Produto", "Resumo Diário", "Resumo Mensal", "Estoque"])

# Função para estilizar cabeçalhos
def style_header(row):
    for cell in row:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

# Criar planilha temporária na memória
def gerar_planilha(dados_venda=None, dados_produto=None):
    wb = Workbook()
    ws_vendas = wb.active
    ws_vendas.title = "Vendas"
    ws_produtos = wb.create_sheet("Produtos")
    ws_resumo_dia = wb.create_sheet("Resumo Diário")
    ws_resumo_mes = wb.create_sheet("Resumo Mensal")
    ws_estoque = wb.create_sheet("Estoque")

    # Cabeçalhos
    headers_vendas = ["Data", "Código Produto", "Produto", "Categoria", "Quantidade",
                      "Preço Unitário (R$)", "Custo Unitário (R$)", "Total Venda (R$)", "Lucro Bruto (R$)"]
    headers_produtos = ["Código", "Produto", "Categoria", "Preço Venda (R$)", "Custo (R$)", "Estoque Atual"]
    headers_resumo_dia = ["Data", "Total Vendas (R$)", "Total Custo (R$)", "Lucro Bruto (R$)", "Quantidade Vendida"]
    headers_resumo_mes = ["Mês", "Total Vendas", "Total Custos", "Lucro Bruto", "Ticket Médio"]
    headers_estoque = ["Código", "Produto", "Estoque Inicial", "Vendido", "Estoque Atual"]

    ws_vendas.append(headers_vendas)
    ws_produtos.append(headers_produtos)
    ws_resumo_dia.append(headers_resumo_dia)
    ws_resumo_mes.append(headers_resumo_mes)
    ws_estoque.append(headers_estoque)

    style_header(ws_vendas[1])
    style_header(ws_produtos[1])
    style_header(ws_resumo_dia[1])
    style_header(ws_resumo_mes[1])
    style_header(ws_estoque[1])

    if dados_venda:
        ws_vendas.append(dados_venda)

    if dados_produto:
        ws_produtos.append(dados_produto)

    # Ajustar largura das colunas
    for sheet in [ws_vendas, ws_produtos, ws_resumo_dia, ws_resumo_mes, ws_estoque]:
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[chr(64 + col)].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# Aba: Cadastro de Venda
if aba == "Cadastrar Venda":
    st.subheader("📦 Cadastrar Venda")
    with st.form("form_venda"):
        col1, col2, col3 = st.columns(3)
        data = col1.date_input("Data da Venda", value=date.today())
        cod_produto = col2.text_input("Código do Produto")
        produto = col3.text_input("Nome do Produto")
        categoria = st.selectbox("Categoria", ["Camisa", "Calça", "Vestido", "Short", "Outro"])
        quantidade = st.number_input("Quantidade", min_value=1, step=1)
        preco_unit = st.number_input("Preço Unitário (R$)", format="%.2f")
        custo_unit = st.number_input("Custo Unitário (R$)", format="%.2f")
        enviado = st.form_submit_button("💾 Gerar Planilha")

        if enviado:
            total = quantidade * preco_unit
            lucro = (preco_unit - custo_unit) * quantidade
            planilha = gerar_planilha([
                data, cod_produto, produto, categoria, quantidade,
                preco_unit, custo_unit, total, lucro
            ])

            st.success("✅ Venda adicionada com sucesso!")
            st.download_button("📥 Baixar Planilha", data=planilha,
                               file_name="Venda_Roupas.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Aba: Cadastro de Produto
elif aba == "Cadastrar Produto":
    st.subheader("🧾 Cadastrar Produto")
    with st.form("form_produto"):
        codigo = st.text_input("Código")
        nome = st.text_input("Nome do Produto")
        categoria = st.text_input("Categoria")
        preco = st.number_input("Preço Venda (R$)", format="%.2f")
        custo = st.number_input("Custo (R$)", format="%.2f")
        estoque = st.number_input("Estoque Atual", min_value=0, step=1)
        enviado = st.form_submit_button("💾 Gerar Planilha")

        if enviado:
            planilha = gerar_planilha(dados_produto=[codigo, nome, categoria, preco, custo, estoque])
            st.success("✅ Produto cadastrado com sucesso!")
            st.download_button("📥 Baixar Planilha", data=planilha,
                               file_name="Produto_Cadastrado.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Abas futuras
else:
    st.info("🚧 Em breve: tela de '{}'".format(aba))
